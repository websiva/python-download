from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from zipfile import BadZipFile
import time
import openpyxl
import csv
import json
import os
from pprint import pprint

file_name = 'input.xlsx'

load_wb = load_workbook(file_name, data_only=True)
load_ws = load_wb['input']
row_count = load_ws.max_row
j = 1
for i in range(2, 21):
    download_directory = 'C:\\Users\\Designer\\Downloads\\New folder\\'+str(j)
    print(download_directory)
    chrome_options = ChromeOptions()
    chrome_options.add_experimental_option('prefs', {
        'download.default_directory': download_directory,
        'download.prompt_for_download': False,
        'download.directory_upgrade': True,
        'safebrowsing.enabled': False,
        'safebrowsing.disable_download_protection': True,
    })


    driver = webdriver.Chrome(
        r"chromedriver_win32\chromedriver.exe", options=chrome_options)

    driver.get("https://chem-prod.treffertech.com/#/product/root")
    driver.maximize_window()
    driver.implicitly_wait(7)
    # ----------------home_category    ----------------------------
    driver.find_element(
        By.XPATH, "//app-product/div/div/div[2]/div[2]/div/div[1]/div/div[2]/a/h4").click()

    # ---------------Sub category ---------------------------------
    driver.implicitly_wait(7)
    driver.find_element(
        By.XPATH, "//app-product/div/div/div[2]/div[2]/div/div[1]/div/div[2]/a/h4").click()
    driver.implicitly_wait(7)
    if row_count >= 4:
        driver.execute_script("window.scrollTo(0,window.scrollY + 400);")
    if row_count >= 9:
        driver.execute_script("window.scrollTo(0,window.scrollY + 800);")
    # ----------------In table click row button -------------------
    driver.implicitly_wait(7)
    driver.find_element(By.XPATH, load_ws['A'+str(i)].value).click()
    driver.implicitly_wait(7)
    #   ---------------create folder-------------
    element = driver.find_element(
        By.XPATH, "/html/body/div[3]/div[2]/div/mat-dialog-container/app-product-dtl-modal/div/div[1]/div[1]/h3")
    download_path = 'C:\\Users\\Designer\\Downloads\\New folder\\'+str(j)
    os.makedirs(download_path)
    # --------------- select CAD selection dropdown-----------------
    driver.find_element(By.XPATH, "//mat-select").click()
    driver.implicitly_wait(7)
    # ---------------select prt option---------------------
    driver.find_element(
        By.XPATH, "//body/div[3]/div[4]/div/div/mat-option[3]").click()
    # -------------- click download button---------------
    driver.implicitly_wait(7)
    driver.find_element(
        By.XPATH, "//body/div[3]/div[2]/div/mat-dialog-container/app-product-dtl-modal/div/div[3]/div[2]/div/button[3]").click()
    driver.implicitly_wait(7)
    # --------------- first name -----------------
    firstname = driver.find_element(
        By.XPATH, "//body/div[3]/div[5]/div/mat-dialog-container/app-rfq/div/div/div/div/div/div[2]/form/div/div/div/input")
    driver.implicitly_wait(7)
    firstname.send_keys('siva')
    # ---------------email id---------------
    firstname = driver.find_element(
        By.XPATH, "//body/div[3]/div[5]/div/mat-dialog-container/app-rfq/div/div/div/div/div/div[2]/form/div[2]/div[2]/div/input")
    driver.implicitly_wait(7)
    firstname.send_keys('sri@gmail.com')
    # ----------------submit button----------
    driver.implicitly_wait(7)
    driver.find_element(
        By.XPATH, "//body/div[3]/div[5]/div/mat-dialog-container/app-rfq/div/div/div/div/div/div[2]/form/button").click()
    wait = WebDriverWait(driver, 60)
    wait.until(lambda x: any(filename.endswith('.zip')
               for filename in os.listdir(download_path)))
    time.sleep(2)
    # --------------- select CAD selection dropdown-----------------
    driver.find_element(By.XPATH, "//mat-select").click()
    driver.implicitly_wait(7)
    # ---------------select prt option---------------------
    driver.find_element(
        By.XPATH, "//body/div[3]/div[4]/div/div/mat-option[5]").click()
    # -------------- click download button---------------
    driver.implicitly_wait(7)
    driver.find_element(
        By.XPATH, "//body/div[3]/div[2]/div/mat-dialog-container/app-product-dtl-modal/div/div[3]/div[2]/div/button[3]").click()
    # --------------- first name -----------------
    firstname = driver.find_element(
        By.XPATH, "//body/div[3]/div[5]/div/mat-dialog-container/app-rfq/div/div/div/div/div/div[2]/form/div/div/div/input")
    driver.implicitly_wait(7)
    firstname.send_keys('siva')
    # ---------------email id---------------
    firstname = driver.find_element(
        By.XPATH, "//body/div[3]/div[5]/div/mat-dialog-container/app-rfq/div/div/div/div/div/div[2]/form/div[2]/div[2]/div/input")
    driver.implicitly_wait(7)
    firstname.send_keys('sri@gmail.com')
    # ----------------submit button----------
    driver.implicitly_wait(7)
    driver.find_element(
        By.XPATH, "//body/div[3]/div[5]/div/mat-dialog-container/app-rfq/div/div/div/div/div/div[2]/form/button").click()
    wait = WebDriverWait(driver, 60)
    wait.until(lambda x: any(filename.endswith('.stp')
               for filename in os.listdir(download_path)))
    time.sleep(2)
    driver.find_element(
        By.XPATH, "//body/div[3]/div/div/mat-dialog-container/app-product-dtl-modal/div/div/div[2]/i").click()
    j += 1
driver.close()
